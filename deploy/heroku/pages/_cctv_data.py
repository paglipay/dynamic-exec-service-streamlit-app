"""
_cctv_data.py — Shared data layer for the camera asset-intake workflow
(Site select -> Upload Asset Workbook + Camera Chart -> Barcode scan ->
Assign -> Export).

Owns exactly two new MongoDB collections, deliberately prefixed so they
can never collide with anything else in the shared database:

    cctv_camera_chart   — the predetermined camera survey (from the
                           uploaded Camera Chart workbook's "Cam Chart"
                           sheet). One doc per (loc_code, camera_id).
    cctv_camera_assets  — scanned/received hardware (from the uploaded
                           Asset Workbook's "New_Equipment" sheet, plus
                           barcode-scanner input). One doc per
                           (loc_code, serial_number).

This module never reads or writes any collection owned by the external
Flask "dynamic-exec-api" service (camera_tasks, camera_mappings,
network_devices, review_queue, submission_CCTV_*, slack_*, sites). It
reads (never writes) the existing `r1_data` site-directory collection
that pages/google_earth.py already uses, for the site picker step.

Camera-ID normalization: three different conventions show up across the
source files —
    Camera Chart "Camera ID(s)":      "1", "2", "38"          (bare int)
    Asset Workbook "AP Number":       "CAM1", "CAM6"           (CAM-prefixed)
    camera_tasks (external) "label":  "01A", "51B"             (2-digit + letter)
normalize_camera_id() reduces all of these to one canonical form so rows
from either uploaded file (and, later, a cross-reference to camera_tasks)
can be matched/joined reliably.
"""

from __future__ import annotations

import io
import re
from typing import Optional

import streamlit as st

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ── Secrets / connection ────────────────────────────────────────────────────

def _get_secret(name: str) -> str:
    """Read from st.secrets if available, fall back to os.environ.
    Matches the convention used by _auth_guard.py, image_cleaner.py,
    google_earth.py, etc."""
    try:
        return str(st.secrets.get(name) or "")
    except Exception:
        import os
        return os.environ.get(name, "")


@st.cache_resource(show_spinner=False)
def _get_client():
    """Return a cached MongoClient, or None if MONGODB_URI isn't set."""
    mongo_uri = _get_secret("MONGODB_URI")
    if not mongo_uri:
        return None
    from pymongo import MongoClient
    return MongoClient(mongo_uri, serverSelectionTimeoutMS=6000)


def get_db():
    """Return the app's Mongo database, or None if not configured/reachable.
    Database name is parsed from the URI path, same convention as
    google_earth.py's _load_r1_schools(), falling back to 'app_data'."""
    mongo_uri = _get_secret("MONGODB_URI")
    if not mongo_uri:
        return None
    client = _get_client()
    if client is None:
        return None
    from urllib.parse import urlparse, unquote
    parsed = urlparse(mongo_uri)
    db_name = unquote((parsed.path or "").lstrip("/")).split("?")[0].strip() or "app_data"
    return client[db_name]


CHART_COLLECTION = "cctv_camera_chart"
ASSETS_COLLECTION = "cctv_camera_assets"


# ── Camera-ID normalization ─────────────────────────────────────────────────

_CAM_ID_RE = re.compile(r"^(?:CAM(?:ERA)?)?\s*0*(\d+)\s*([A-Za-z]?)\s*$", re.IGNORECASE)


def normalize_camera_id(raw) -> Optional[dict]:
    """Parse any of '1', 'CAM6', '06', '51B', 'CAM06A' into
    {'num': int, 'letter': str, 'canonical': str}, or None if unparseable.
    'canonical' is zero-padded 2-digit + optional letter, e.g. '06', '51B' —
    the same shape camera_tasks (external) uses for its 'label' field, so a
    future read-only cross-reference is a straight string match."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.upper() in {"#N/A", "N/A", "NA"}:
        return None
    m = _CAM_ID_RE.match(s)
    if not m:
        return None
    num = int(m.group(1))
    letter = m.group(2).upper()
    return {"num": num, "letter": letter, "canonical": f"{num:02d}{letter}"}


def extract_camera_id_from_text(text) -> Optional[dict]:
    """Some Asset Workbook rows embed the camera number inside a longer
    string, e.g. Cabinet = 'CLDF-G3-CAM06'. Find the first CAM<n> token
    anywhere in the text and normalize it."""
    if not text:
        return None
    m = re.search(r"CAM\s*0*(\d+)\s*([A-Za-z]?)", str(text), re.IGNORECASE)
    if not m:
        return None
    return normalize_camera_id(f"CAM{m.group(1)}{m.group(2)}")


# ── Excel parsing ────────────────────────────────────────────────────────────

def _find_header_row(ws, must_contain: str, max_scan_rows: int = 10):
    """Return the 1-based row index of the first row whose first few cells
    contain `must_contain` (case-insensitive), scanning up to
    max_scan_rows. Falls back to row 1 if not found."""
    needle = must_contain.strip().lower()
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, max_col=min(ws.max_column, 25), values_only=True),
        start=1,
    ):
        for cell in row:
            if cell and needle in str(cell).strip().lower():
                return i
    return 1


def _rows_as_dicts(ws, header_row: int):
    """Yield {header: value} dicts for every non-empty row after header_row."""
    headers = [
        (str(h).strip() if h is not None else "")
        for h in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    ]
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        rec = {}
        for h, v in zip(headers, row):
            if h:
                rec[h] = v
        if rec:
            yield rec


def parse_asset_workbook(file_bytes: bytes, sheet_name: str = "New_Equipment") -> list[dict]:
    """Parse the Asset Workbook's populated asset-row sheet (New_Equipment
    by default — confirmed as the authoritative source; Assets_T_and_A is
    a blank hand-off template, newinstallprepsheet duplicates the same
    rows with a few extra install-phase fields).

    Returns a list of normalized dicts ready for cctv_camera_assets:
        loc_code, site_name, building, floor, room_number, cafm_room_number,
        cabinet, camera_id (normalized dict or None), camera_id_raw,
        manufacturer, model_number, serial_number, mac_address, ip_address,
        host_name, component_of, equipment_category, po_number,
        contract_number, project, notes, source_row (raw dict, for audit)
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to parse .xlsm files")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True, keep_links=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    header_row = _find_header_row(ws, "Manufacturer")

    out = []
    for rec in _rows_as_dicts(ws, header_row):
        serial = str(rec.get("Serial Number") or "").strip()
        if not serial or serial.upper() in {"#N/A", "N/A"}:
            continue  # skip template/blank rows

        ap_number = rec.get("AP Number")
        cam = normalize_camera_id(ap_number) or extract_camera_id_from_text(rec.get("Cabinet"))

        out.append({
            "loc_code": str(rec.get("Location Code") or "").strip(),
            "site_name": str(rec.get("Site Name") or "").strip(),
            "building": rec.get("Building"),
            "floor": rec.get("Floor"),
            "room_number": rec.get("Room Number"),
            "cafm_room_number": rec.get("CAFM Room Number"),
            "cabinet": rec.get("Cabinet"),
            "camera_id_raw": ap_number,
            "camera_id": cam,
            "manufacturer": rec.get("Manufacturer"),
            "model_number": str(rec.get("Model Number") or rec.get("Model Number ") or "").strip(),
            "serial_number": serial,
            "mac_address": str(rec.get("MAC Address") or "").strip() or None,
            "ip_address": str(rec.get("IP Address") or "").strip() or None,
            "host_name": rec.get("Host Name (DNS Name)"),
            "component_of": rec.get("Component Of"),
            "equipment_category": rec.get("Equipment Category"),
            "po_number": rec.get("PO Number"),
            "contract_number": rec.get("Contract Number"),
            "project": rec.get("Project"),
            "notes": rec.get("Notes"),
            "status": "received",
            "source_row": rec,
        })
    return out


def parse_camera_chart(file_bytes: bytes, sheet_name: str = "Cam Chart") -> list[dict]:
    """Parse the Camera Chart workbook's camera-survey sheet.

    Returns a list of normalized dicts ready for cctv_camera_chart:
        camera_id (normalized dict), camera_id_raw, building,
        cafm_building_id, floor, indoor_outdoor, mount_type, mount_height,
        view_description, camera_model_text, num_views, data_cabinet,
        cabinet_room_location, switch_name, cable_label_id, network_notes,
        source_row (raw dict, for audit)
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to parse .xlsm files")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True, keep_links=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Site name lives a row or two above the header, e.g. "Site: State
    # Street Elementary (6918)" — pull it for cross-check against the
    # selected site, independent of finding the header row itself.
    site_line = ""
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=6, values_only=True):
        for i, cell in enumerate(row):
            if cell and str(cell).strip().lower().startswith("site"):
                rest = [c for c in row[i + 1:] if c]
                if rest:
                    site_line = str(rest[0]).strip()
                break
        if site_line:
            break

    header_row = _find_header_row(ws, "Camera ID")

    out = []
    for rec in _rows_as_dicts(ws, header_row):
        raw_id = rec.get("Camera ID(s)") or rec.get("Camera ID")
        cam = normalize_camera_id(raw_id)
        if cam is None:
            continue  # skip notes/blank rows that slipped past the emptiness check

        out.append({
            "camera_id_raw": raw_id,
            "camera_id": cam,
            "building": rec.get("Building"),
            "cafm_building_id": rec.get("CAFM Building ID"),
            "cafm_space_id": rec.get("CAFM Space ID"),
            "floor": rec.get("Floor"),
            "indoor_outdoor": rec.get("Indoor or Outdoor"),
            "mount_type": rec.get("Mount Type"),
            "mount_height": rec.get("Mount Hgt (Aprox)"),
            "view_description": rec.get("View Description"),
            "camera_model_text": rec.get("Camera Model"),
            "num_views": rec.get("Num of Views"),
            "data_cabinet": rec.get("Data Cabinet"),
            "cabinet_room_location": rec.get("Cabinet Room Location"),
            "switch_name": rec.get("Switch Name"),
            "cable_label_id": rec.get("Cable Label ID"),
            "network_notes": rec.get("Network Notes"),
            "status": "planned",
            "source_row": rec,
        })
    return out, site_line


# ── Mongo I/O ────────────────────────────────────────────────────────────────

def upsert_camera_chart_rows(loc_code: str, rows: list[dict]) -> int:
    """Upsert parsed Camera Chart rows, keyed by (loc_code, camera_id.canonical).
    Returns the number of rows written."""
    db = get_db()
    if db is None:
        raise RuntimeError("MONGODB_URI is not configured")
    coll = db[CHART_COLLECTION]
    n = 0
    for row in rows:
        cam = row.get("camera_id")
        if not cam:
            continue
        doc = dict(row)
        doc["loc_code"] = loc_code
        coll.update_one(
            {"loc_code": loc_code, "camera_id.canonical": cam["canonical"]},
            {"$set": doc},
            upsert=True,
        )
        n += 1
    return n


def upsert_camera_assets_rows(loc_code: str, rows: list[dict]) -> int:
    """Upsert parsed/scanned asset rows, keyed by (loc_code, serial_number).
    Returns the number of rows written."""
    db = get_db()
    if db is None:
        raise RuntimeError("MONGODB_URI is not configured")
    coll = db[ASSETS_COLLECTION]
    n = 0
    for row in rows:
        serial = row.get("serial_number")
        if not serial:
            continue
        doc = dict(row)
        doc["loc_code"] = loc_code
        coll.update_one(
            {"loc_code": loc_code, "serial_number": serial},
            {"$set": doc},
            upsert=True,
        )
        n += 1
    return n


def get_camera_chart(loc_code: str) -> list[dict]:
    db = get_db()
    if db is None:
        return []
    return list(db[CHART_COLLECTION].find({"loc_code": loc_code}, {"_id": 0}).sort("camera_id.num", 1))


def get_camera_assets(loc_code: str) -> list[dict]:
    db = get_db()
    if db is None:
        return []
    return list(db[ASSETS_COLLECTION].find({"loc_code": loc_code}, {"_id": 0}))


# ── Matching (camera-ID assignment) ─────────────────────────────────────────
#
# UNVALIDATED: this heuristic has only been checked against two files from
# *different* sites (parsing correctness only). The actual assignment rule
# — which open Camera Chart slot a given scanned unit should claim — needs
# confirming against a real matched Camera-Chart + Asset-Workbook pair from
# one project before this is trusted for real intake. Every suggestion here
# must be reviewed by a human before being confirmed (see
# camera_assign_review.py) — nothing auto-commits.

def suggest_assignments(loc_code: str) -> list[dict]:
    """For every unassigned scanned asset at this site, suggest the lowest
    open Camera Chart camera_id whose camera_model_text contains the
    asset's model_number as a substring (case-insensitive) — the pattern
    observed in the sample files, e.g. asset Model Number 'P3748-PLVE'
    inside chart Camera Model text 'AXIS P3748-PLVE 4 Element 270-360...'.

    Returns [{"asset": {...}, "candidates": [chart_row, ...]}, ...] —
    candidates sorted lowest camera_id first, already-claimed-by-an-
    earlier-suggestion-in-this-batch slots excluded so two scanned units
    of the same model don't both get suggested the same slot.
    """
    chart = [r for r in get_camera_chart(loc_code) if r.get("status") == "planned"]
    assets = [r for r in get_camera_assets(loc_code) if not r.get("camera_id") and r.get("status") == "received"]

    claimed_canonicals: set[str] = set()
    out = []
    for asset in assets:
        model = (asset.get("model_number") or "").strip().upper()
        candidates = []
        if model:
            for row in chart:
                canon = row["camera_id"]["canonical"]
                if canon in claimed_canonicals:
                    continue
                text = (row.get("camera_model_text") or "").upper()
                if model in text:
                    candidates.append(row)
            candidates.sort(key=lambda r: (r["camera_id"]["num"], r["camera_id"]["letter"]))
        if candidates:
            claimed_canonicals.add(candidates[0]["camera_id"]["canonical"])
        out.append({"asset": asset, "candidates": candidates})
    return out


def confirm_assignment(loc_code: str, serial_number: str, camera_id_canonical: str) -> bool:
    """Attach a Camera Chart slot to a scanned asset. Marks the chart
    slot 'assigned' and stamps the asset with that camera_id + status
    'assigned'. Returns False if either side is already taken.

    Atomic by construction, not check-then-set: with multiple sessions
    (multiple techs/desks) potentially confirming against the same
    site's chart concurrently — e.g. two auto_assign_on_scan() calls
    landing at the same instant — a plain find_one() followed by a
    separate update_one() has a race where both callers could pass the
    check before either writes. Each step below folds its own
    precondition ('still planned' / 'still unassigned') into the
    update's filter, so MongoDB itself guarantees only one caller wins
    a given slot.
    """
    db = get_db()
    if db is None:
        raise RuntimeError("MONGODB_URI is not configured")

    # find_one_and_update's default returns the doc as it was BEFORE
    # this update -- camera_id itself is never touched here, so that's
    # fine to read off it below.
    chart_row = db[CHART_COLLECTION].find_one_and_update(
        {"loc_code": loc_code, "camera_id.canonical": camera_id_canonical, "status": "planned"},
        {"$set": {"status": "assigned", "assigned_serial_number": serial_number}},
    )
    if not chart_row:
        return False  # slot doesn't exist, or another caller just claimed it

    asset_result = db[ASSETS_COLLECTION].update_one(
        {"loc_code": loc_code, "serial_number": serial_number, "camera_id": None},
        {"$set": {"camera_id": chart_row["camera_id"], "status": "assigned"}},
    )
    if asset_result.modified_count == 0:
        # Asset doesn't exist, or another caller already assigned it —
        # release the chart slot we just claimed so it's not stuck
        # orphaned as "assigned" with no asset actually attached.
        db[CHART_COLLECTION].update_one(
            {"loc_code": loc_code, "camera_id.canonical": camera_id_canonical},
            {"$set": {"status": "planned"}, "$unset": {"assigned_serial_number": ""}},
        )
        return False

    return True


def add_scanned_asset(loc_code: str, serial_number: str, model_number: str, **extra) -> None:
    """Record one barcode scan as a new (or updated) received asset,
    unassigned until confirm_assignment() runs. `extra` may include
    manufacturer, mac_address, ip_address, notes, etc."""
    row = {
        "serial_number": str(serial_number).strip(),
        "model_number": str(model_number).strip(),
        "camera_id": None,
        "camera_id_raw": None,
        "status": "received",
        **extra,
    }
    upsert_camera_assets_rows(loc_code, [row])


def auto_assign_on_scan(loc_code: str, serial_number: str) -> Optional[dict]:
    """Scan-time auto-assign: immediately claim the lowest open Camera
    Chart slot whose model matches this one asset — no human review
    pause, unlike suggest_assignments()/camera_assign_review.py's
    reviewed flow. Deliberately picks the lowest candidate even when
    several open slots share the model (see camera_assign_review.py's
    docstring for why that's ambiguous — a chosen tradeoff, not an
    oversight, made so a label can print right after each scan).

    Returns the assigned camera_id dict, or None if there's no chart
    uploaded yet, the asset doesn't exist/is already assigned, or no
    open slot matches this asset's model.
    """
    db = get_db()
    if db is None:
        return None

    asset = db[ASSETS_COLLECTION].find_one({"loc_code": loc_code, "serial_number": serial_number})
    if not asset or asset.get("camera_id"):
        return None

    model = (asset.get("model_number") or "").strip().upper()
    if not model:
        return None

    candidates = list(db[CHART_COLLECTION].find({"loc_code": loc_code, "status": "planned"}))
    matches = [c for c in candidates if model in (c.get("camera_model_text") or "").upper()]
    if not matches:
        return None
    matches.sort(key=lambda r: (r["camera_id"]["num"], r["camera_id"]["letter"]))
    chosen = matches[0]

    ok = confirm_assignment(loc_code, serial_number, chosen["camera_id"]["canonical"])
    return chosen["camera_id"] if ok else None


# ── Print broker (slack-to-onedrive-sync's /print-jobs; see that repo's
# app.py + sync_lib.py "Camera-label print queue" section) ─────────────────
#
# Multi-desk routing: every print agent has its own persistent device_id
# (see local_print_agent/agent_config.py), so a job must say which one
# should print it — list_print_devices() lists agents that have polled
# recently, for the session's device picker (see camera_site_select.py),
# and its choice is passed into enqueue_print_job() as device_id. Without
# this, every running agent would grab every job regardless of site/desk,
# duplicate-printing across every printer in the building.

def list_print_devices() -> list[dict]:
    """Print agents that have polled the broker recently (active print
    agents), for the session's device picker. Returns [] on any broker
    problem — the picker just shows nothing rather than erroring, since
    this is polled from the UI on every rerun."""
    broker_url = _get_secret("PRINT_BROKER_URL")
    broker_secret = _get_secret("PRINT_BROKER_SECRET")
    if not broker_url or not broker_secret:
        return []
    try:
        import requests
        resp = requests.get(
            f"{broker_url.rstrip('/')}/print-jobs/devices",
            headers={"Authorization": f"Bearer {broker_secret}"},
            timeout=5,
        )
        if resp.status_code == 200:
            return resp.json().get("items", [])
        return []
    except Exception:
        return []


def enqueue_print_job(site_name: str, loc_code: str, camera_id: dict, serial_number: str,
                       model_number: str, device_id: str) -> dict:
    """Best-effort POST to the broker, targeting one specific print
    agent (device_id — see list_print_devices()). Never raises — a
    broker outage shouldn't break the scan loop, just skip printing.
    Returns {"ok": bool, "error": str | None}."""
    broker_url = _get_secret("PRINT_BROKER_URL")
    broker_secret = _get_secret("PRINT_BROKER_SECRET")
    if not broker_url or not broker_secret:
        return {"ok": False, "error": "PRINT_BROKER_URL/PRINT_BROKER_SECRET not configured"}
    if not device_id:
        return {"ok": False, "error": "No print device selected"}

    try:
        import requests
        camera_number = f"CAM{camera_id['num']}{camera_id['letter']}"
        resp = requests.post(
            f"{broker_url.rstrip('/')}/print-jobs",
            headers={"Authorization": f"Bearer {broker_secret}"},
            json={
                "camera_number": camera_number,
                "serial_number": serial_number,
                "model_number": model_number,
                "site_name": site_name,
                "loc_code": loc_code,
                "device_id": device_id,
            },
            timeout=5,
        )
        if resp.status_code == 201:
            return {"ok": True, "error": None}
        return {"ok": False, "error": f"HTTP {resp.status_code}: {resp.text[:200]}"}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


# ── Site directory (read-only; owned by google_earth.py's r1_data) ─────────

@st.cache_data(ttl=300, show_spinner=False)
def list_sites() -> list[dict]:
    """Read-only read of the existing r1_data site directory. Never writes
    here — google_earth.py owns this collection."""
    db = get_db()
    if db is None:
        return []
    cursor = db["r1_data"].find(
        {},
        {"_id": 0, "School Name": 1, "Site": 1, "Loc Code": 1, "Address": 1, "City": 1, "Contractor": 1},
    ).sort("School Name", 1)
    return [dict(doc) for doc in cursor]


# ── Barcode image decoding + classification (camera_barcode_settings.py) ───
#
# Global, not per-site: barcode format is a property of the hardware line
# (Axis), not the school. Rules are ordered; the first regex that matches
# a decoded barcode's text wins. A catch-all ".*" rule is required as the
# last entry so every decoded value always classifies as something —
# get_barcode_rules() re-appends one if it's ever missing (e.g. someone
# deleted every row in the settings page).
#
# `field` is one of "serial_number", "model_number", or "ignore" — ignore
# is for barcodes that are neither, e.g. the EAN-13 retail/GTIN barcode
# printed on most Axis boxes alongside the actual Part No./Serial No.
# barcodes (confirmed against a real AXIS T90A21 IR-LED box label: three
# barcodes present — Part No. 5013-211, Serial No. HW013509110, and a
# 13-digit EAN-13 — the last of which must be dropped, not counted as a
# second "model", or it blocks the "exactly one Model + one Serial"
# auto-add check in camera_barcode_scan.py).
#
# B8A4-prefixed serials are MAC-derived (B8:A4:4F is Axis's registered
# MAC OUI), which only applies to networked devices whose serial IS their
# MAC address (network cameras). Non-networked accessories (like that
# IR-LED illuminator) get a different serial format — HW-prefixed, per
# that same real label — hence two separate serial rules below rather
# than assuming one prefix covers every Axis product.

BARCODE_RULES_COLLECTION = "cctv_barcode_rules"
BARCODE_RULES_DOC_ID = "default"  # single global doc; not per-site

DEFAULT_BARCODE_RULES = [
    {"pattern": "^B8A4", "field": "serial_number", "label": "Serial Number (MAC-derived, networked devices)", "case_insensitive": True},
    {"pattern": "^HW", "field": "serial_number", "label": "Serial Number (HW-prefixed, accessories)", "case_insensitive": True},
    {"pattern": r"^\d{13}$", "field": "ignore", "label": "EAN-13 retail barcode (ignored)", "case_insensitive": True},
    {"pattern": ".*", "field": "model_number", "label": "Model Number (default)", "case_insensitive": True},
]


def get_barcode_rules() -> list[dict]:
    """Ordered classification rules, seeding the default set on first use."""
    db = get_db()
    if db is None:
        return list(DEFAULT_BARCODE_RULES)
    doc = db[BARCODE_RULES_COLLECTION].find_one({"_id": BARCODE_RULES_DOC_ID})
    rules = (doc or {}).get("rules")
    if not rules:
        rules = list(DEFAULT_BARCODE_RULES)
        save_barcode_rules(rules)
    # Guard against a saved set that's missing a catch-all (e.g. every rule
    # was deleted in the settings page) — classify_barcode() depends on one
    # always matching.
    if not any(r.get("pattern") == ".*" for r in rules):
        rules = rules + [{"pattern": ".*", "field": "model_number", "label": "Model Number (default)", "case_insensitive": True}]
    return rules


def save_barcode_rules(rules: list[dict]) -> None:
    db = get_db()
    if db is None:
        raise RuntimeError("MONGODB_URI is not configured")
    db[BARCODE_RULES_COLLECTION].update_one(
        {"_id": BARCODE_RULES_DOC_ID},
        {"$set": {"rules": rules}},
        upsert=True,
    )


def classify_barcode(text: str, rules: Optional[list[dict]] = None) -> Optional[str]:
    """Return the target field ('serial_number' or 'model_number') for a
    decoded barcode string, per the first matching rule. None only if
    `rules` has no catch-all (get_barcode_rules() always adds one; a
    caller passing a hand-built list without one can still get None)."""
    if rules is None:
        rules = get_barcode_rules()
    for rule in rules:
        pattern = rule.get("pattern") or ""
        flags = re.IGNORECASE if rule.get("case_insensitive", True) else 0
        try:
            if re.search(pattern, text, flags):
                return rule.get("field")
        except re.error:
            continue  # a bad regex in a saved rule shouldn't break every scan
    return None


def decode_barcodes_from_image(image_bytes: bytes) -> list[str]:
    """Decode every barcode found in an image (any symbology zxing-cpp
    supports — Code128, Code39, EAN, QR, DataMatrix, etc.), e.g. a single
    photo of a label carrying both the Model and Serial barcodes.
    Returns decoded text strings, empty list if none found or on any
    decode failure (never raises — a bad/corrupt image shouldn't break
    the scan page)."""
    try:
        import zxingcpp
        from PIL import Image
        img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
        results = zxingcpp.read_barcodes(img)
        return [r.text for r in results if r.valid and r.text]
    except Exception:
        return []


# ── OCR text reading (fallback/complement to barcode decoding) ─────────────
#
# rapidocr-onnxruntime, not pytesseract: pytesseract needs the Tesseract
# binary installed at the OS level, which Heroku's default Python
# buildpack doesn't provide without extra buildpack/Aptfile work. rapidocr
# is a pure pip package with its own bundled ONNX models (~15MB) and
# reuses onnxruntime, already a dependency here (image_cleaner.py's YOLO
# detection) — no new heavy runtime, no system binary.
#
# Real deployment gotcha, already handled: rapidocr-onnxruntime hard-
# depends on GUI `opencv-python` (confirmed against PyPI metadata, no
# headless-friendly release exists), which conflicts with this app's
# `opencv-python-headless` (also image_cleaner.py) — both packages share
# the same `cv2/` install directory and don't coexist cleanly (confirmed
# locally: having both installed, even briefly, can leave `cv2` broken —
# e.g. `cv2.resize` missing — even after uninstalling the GUI one again).
# See bin/post_compile, which swaps GUI opencv-python back out for
# opencv-python-headless right after `pip install -r requirements.txt` —
# a documented Heroku Python buildpack hook, not custom infra.

_MODEL_LABEL_RE = re.compile(r"(?:part\s*no\.?|model\s*no\.?|model)[:\s]*([A-Z0-9][A-Z0-9\-/]{2,})", re.IGNORECASE)
# Second model pattern: on some labels (e.g. AXIS Q1786-LE) the model is
# printed straight after "AXIS" with no "Model:"/"Part No." caption at
# all -- confirmed against a real label read this way. Requires a digit
# in the captured value so it doesn't also match "AXIS COMMUNICATIONS"
# (the brand's own logo text, which has no digit) as a fake "model".
_MODEL_AXIS_RE = re.compile(r"\bAXIS[:\s]*([A-Z][A-Z0-9]*\d[A-Z0-9\-]{2,})", re.IGNORECASE)
_SERIAL_LABEL_RE = re.compile(r"(?:serial\s*no\.?|s\s*/\s*n)[:\s]*([A-Z0-9][A-Z0-9\-/]{2,})", re.IGNORECASE)

_ocr_engine = None


def _get_ocr_engine():
    """Lazily construct and cache the OCR engine — it loads ONNX models
    from disk on first use, so this avoids paying that cost until an
    image is actually uploaded."""
    global _ocr_engine
    if _ocr_engine is None:
        from rapidocr_onnxruntime import RapidOCR
        _ocr_engine = RapidOCR()
    return _ocr_engine


def read_text_from_image(image_bytes: bytes) -> list[str]:
    """OCR every text line found in an image. Returns [] on any failure
    (missing engine, corrupt image, no text found) — never raises, same
    convention as decode_barcodes_from_image()."""
    try:
        import numpy as np
        from PIL import Image
        img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
        engine = _get_ocr_engine()
        result, _elapse = engine(np.array(img))
        if not result:
            return []
        return [text for _box, text, _score in result if text]
    except Exception:
        return []


def extract_model_serial_from_text(lines: list[str], rules: Optional[list[dict]] = None) -> dict:
    """Best-effort: find Model/Serial values in OCR'd label text lines.

    Two passes, because a real AXIS label's layout isn't consistent
    (confirmed against a real T90A21 box label): "Part No. 5013-211" is
    one line — label and value together — but "Serial No." is its own
    line, with the actual value ("HW013509110") printed separately below
    the barcode, so OCR returns them as two unrelated text lines with
    nothing connecting them.

    1. Explicit "Part No./Model" and "Serial No./S/N" label prefixes,
       captured from whatever follows on the SAME line/text region.
    2. If no "Model"/"Part No." caption was found, try the "AXIS
       <model>" pattern instead (_MODEL_AXIS_RE) — some labels (e.g. a
       real AXIS Q1786-LE box) print the model straight after the brand
       name with no caption word at all.
    3. If no explicit serial label+value was found on one line, fall
       back to running the barcode classification rules (same ones
       classify_barcode() uses — B8A4/HW prefixes etc.) against every
       OCR'd line, since a bare serial number is usually distinctive
       enough on its own even without its caption attached. This
       fallback is serial-only (not model) because the model catch-all
       rule (".*") is too permissive for free OCR text — it would treat
       company names, certification marks, etc. as a "model" too.

    Returns {"model": str|None, "serial": str|None} — either or both
    may be None if not found.
    """
    model = None
    serial = None
    joined = " ".join(lines)

    m = _MODEL_LABEL_RE.search(joined)
    if m:
        model = m.group(1).strip()
    if not model:
        m = _MODEL_AXIS_RE.search(joined)
        if m:
            model = m.group(1).strip()

    m = _SERIAL_LABEL_RE.search(joined)
    if m:
        serial = m.group(1).strip()

    if not serial:
        for line in lines:
            cleaned = line.strip()
            if cleaned and classify_barcode(cleaned, rules) == "serial_number":
                serial = cleaned
                break

    return {"model": model, "serial": serial}
